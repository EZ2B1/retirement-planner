"""Tests for IRMAA projection, workbook, and CSV reporting."""

from __future__ import annotations

import csv

from openpyxl import load_workbook

from retirement_planner.calculators.irmaa import calculate_irmaa
from retirement_planner.calculators.irmaa_strategy import (
    IrmaaIncomeSource,
    analyze_irmaa_income_sources,
    evaluate_irmaa_strategy,
)
from retirement_planner.engines.irmaa_projection import project_irmaa_year
from retirement_planner.reporting.excel_builder import REQUIRED_WORKSHEET_ORDER
from retirement_planner.reporting.irmaa_reporting import (
    build_irmaa_validation_candidate,
    export_irmaa_audit_csv,
)


def _projection():
    sources = (
        IrmaaIncomeSource("Pension and taxable income", 210000.0),
        IrmaaIncomeSource("Roth conversion", 12000.0),
    )
    diagnostic = analyze_irmaa_income_sources(
        income_year=2024,
        filing_status="married_filing_jointly",
        beneficiary_count=2,
        sources=sources,
    )
    strategy = evaluate_irmaa_strategy(
        income_year=2024,
        filing_status="married_filing_jointly",
        beneficiary_count=2,
        baseline_agi=210000.0,
        baseline_tax_exempt_interest=0.0,
        added_agi=12000.0,
    )
    baseline = calculate_irmaa(
        income_year=2024,
        filing_status="married_filing_jointly",
        agi=210000.0,
        tax_exempt_interest=0.0,
        beneficiary_count=2,
    )
    return project_irmaa_year(
        premium_result=baseline,
        strategy_decision=strategy,
    ), diagnostic


def test_projection_rejects_nonreconciling_diagnostic() -> None:
    projection, diagnostic = _projection()
    try:
        project_irmaa_year(
            premium_result=projection.premium_result,
            income_diagnostic=diagnostic,
        )
    except ValueError as exc:
        assert "does not reconcile" in str(exc)
    else:
        raise AssertionError("expected reconciliation failure")


def test_validation_candidate_preserves_sheet_order_and_reports_warning(tmp_path) -> None:
    projection, _ = _projection()
    output = tmp_path / "IRMAA_Validation_Candidate.xlsx"

    build_irmaa_validation_candidate(projection, output)
    workbook = load_workbook(output, data_only=False)

    assert tuple(workbook.sheetnames) == REQUIRED_WORKSHEET_ORDER
    assert workbook["Medicare IRMAA"]["A1"].value.endswith("Validation Candidate")
    assert workbook["Medicare IRMAA"]["B16"].value is True
    assert workbook["Medicare IRMAA"]["B17"].value is False
    assert "crosses an IRMAA threshold" in workbook["Advisor Report"]["A3"].value
    assert "Do not implement" in workbook["Recommendations"]["A2"].value
    assert workbook["Validation"]["E2"].value == "PASS"


def test_irmaa_csv_exports_reconciled_household_fields(tmp_path) -> None:
    projection, _ = _projection()
    output = tmp_path / "irmaa_audit.csv"

    export_irmaa_audit_csv(projection, output)
    with output.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))

    assert len(rows) == 1
    assert rows[0]["premium_year"] == "2026"
    assert rows[0]["filing_status"] == "married_filing_jointly"
    assert rows[0]["beneficiary_count"] == "2"
    assert rows[0]["threshold_warning"]


def test_projection_rejects_wrong_premium_year() -> None:
    result = calculate_irmaa(
        income_year=2024,
        filing_status="single",
        agi=100000.0,
        tax_exempt_interest=0.0,
        beneficiary_count=1,
    )
    projection = project_irmaa_year(premium_result=result)
    object.__setattr__(projection, "projection_year", 2025)

    try:
        projection.validate_reconciliation()
    except ValueError as exc:
        assert "projection year" in str(exc)
    else:
        raise AssertionError("expected premium-year reconciliation failure")
