"""Tests for workbook structure and RMD validation-candidate reporting."""

from datetime import date

from openpyxl import load_workbook

from retirement_planner.calculators.rmd_compliance import RmdAccountInput
from retirement_planner.engines.projection_engine import project_rmd_year
from retirement_planner.reporting.excel_builder import (
    REQUIRED_WORKSHEET_ORDER,
    RESTORED_REPORTING_SHEETS,
    build_rmd_validation_candidate,
    create_workbook_shell,
    populate_rmd_analysis,
    validate_sheet_structure,
)


def _projection_result():
    return project_rmd_year(
        2026,
        (
            RmdAccountInput(
                account_id="IRA-1",
                owner_id="owner-1",
                account_type="traditional_ira",
                prior_year_end_balance=274000.0,
                divisor=27.4,
                life_expectancy_table="Uniform Lifetime",
                distributions_taken=4000.0,
                qcd_requested=6000.0,
                qcd_direct_transfer=True,
                owner_date_of_birth=date(1950, 1, 1),
                qcd_distribution_date=date(2026, 2, 1),
            ),
        ),
        qcd_annual_limit=108000.0,
    )


def test_workbook_shell_uses_required_user_workflow_order() -> None:
    workbook = create_workbook_shell()
    assert tuple(workbook.sheetnames) == REQUIRED_WORKSHEET_ORDER
    assert validate_sheet_structure(workbook.sheetnames) == []


def test_manifest_contains_all_issue_1_reporting_sheets() -> None:
    assert RESTORED_REPORTING_SHEETS.issubset(set(REQUIRED_WORKSHEET_ORDER))


def test_validation_reports_missing_required_sheet() -> None:
    sheet_names = [name for name in REQUIRED_WORKSHEET_ORDER if name != "Scenario Comparison"]
    errors = validate_sheet_structure(sheet_names)
    assert any("Scenario Comparison" in error for error in errors)


def test_validation_reports_misordered_required_sheets() -> None:
    sheet_names = list(REQUIRED_WORKSHEET_ORDER)
    sheet_names[0], sheet_names[1] = sheet_names[1], sheet_names[0]
    errors = validate_sheet_structure(sheet_names)
    assert "Required worksheets are not in the approved user-workflow order" in errors


def test_validation_allows_nonrequired_support_sheet_without_reordering() -> None:
    sheet_names = list(REQUIRED_WORKSHEET_ORDER)
    sheet_names.append("Temporary QA Notes")
    assert validate_sheet_structure(sheet_names) == []


def test_populate_rmd_analysis_writes_reconciled_values() -> None:
    workbook = create_workbook_shell()
    result = _projection_result()
    populate_rmd_analysis(workbook, result)

    rmd = workbook["RMD Analysis"]
    assert rmd["B2"].value == 2026
    assert rmd["B3"].value == "COMPLIANT"
    assert rmd["B5"].value == 10000.0
    assert rmd["B6"].value == 6000.0
    assert rmd["B7"].value == 4000.0
    assert rmd["B8"].value == 0.0
    assert rmd["B12"].value == "IRA-1"
    assert workbook["Validation"]["B2"].value == "PASS"


def test_build_rmd_validation_candidate_preserves_baseline_separation(tmp_path) -> None:
    output = tmp_path / "validation" / "RMD_Validation_Candidate.xlsx"
    built = build_rmd_validation_candidate(_projection_result(), output)

    assert built == output
    assert output.exists()
    workbook = load_workbook(output, data_only=False)
    assert tuple(workbook.sheetnames) == REQUIRED_WORKSHEET_ORDER
    assert workbook["RMD Analysis"]["A1"].value.endswith("Validation Candidate")
    assert workbook["Audit"]["B3"].value == 1


def test_validation_candidate_requires_xlsx_extension(tmp_path) -> None:
    try:
        build_rmd_validation_candidate(_projection_result(), tmp_path / "candidate.xls")
    except ValueError as exc:
        assert ".xlsx" in str(exc)
    else:
        raise AssertionError("non-xlsx validation candidate should be rejected")
