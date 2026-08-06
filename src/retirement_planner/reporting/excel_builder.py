"""Excel workbook structure and RMD validation-candidate reporting.

This module defines the required user-facing worksheet order for the formula-driven
validation candidate. It deliberately does not modify or replace the approved base
workbook.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font

from retirement_planner.models.results import RmdAnnualProjectionResult


@dataclass(frozen=True)
class WorksheetDefinition:
    """Metadata for a required worksheet."""

    name: str
    section: str
    required: bool = True


WORKSHEET_MANIFEST: tuple[WorksheetDefinition, ...] = (
    WorksheetDefinition("Cover", "Orientation and control"),
    WorksheetDefinition("Instructions", "Orientation and control"),
    WorksheetDefinition("Dashboard", "Orientation and control"),
    WorksheetDefinition("Scenario Manager", "Orientation and control"),
    WorksheetDefinition("Household Inputs", "Household and planning inputs"),
    WorksheetDefinition("Economic Assumptions", "Household and planning inputs"),
    WorksheetDefinition("Asset-Class Assumptions", "Household and planning inputs"),
    WorksheetDefinition("Tax Assumptions", "Household and planning inputs"),
    WorksheetDefinition("Social Security", "Household and planning inputs"),
    WorksheetDefinition("Pension and Other Income", "Household and planning inputs"),
    WorksheetDefinition("Account Balances", "Household and planning inputs"),
    WorksheetDefinition("Spending and Cash Needs", "Household and planning inputs"),
    WorksheetDefinition("Annual Projection", "Core annual calculations"),
    WorksheetDefinition("Lifetime Cash Flow", "Core annual calculations"),
    WorksheetDefinition("Federal Tax Calculation", "Core annual calculations"),
    WorksheetDefinition("RMD Analysis", "Core annual calculations"),
    WorksheetDefinition("Medicare IRMAA", "Core annual calculations"),
    WorksheetDefinition("ACA Analysis", "Core annual calculations"),
    WorksheetDefinition("Healthcare Planning", "Core annual calculations"),
    WorksheetDefinition("Roth Conversion Optimizer", "Strategy and optimization"),
    WorksheetDefinition("Tax Optimizer", "Strategy and optimization"),
    WorksheetDefinition("Withdrawal Strategy", "Strategy and optimization"),
    WorksheetDefinition("Social Security Claiming", "Strategy and optimization"),
    WorksheetDefinition("Charitable Planning", "Strategy and optimization"),
    WorksheetDefinition("Estate Planning", "Strategy and optimization"),
    WorksheetDefinition("Scenario Comparison", "Comparison and risk analysis"),
    WorksheetDefinition("Sensitivity Analysis", "Comparison and risk analysis"),
    WorksheetDefinition("Monte Carlo Analysis", "Comparison and risk analysis"),
    WorksheetDefinition("Optimizer Results", "Comparison and risk analysis"),
    WorksheetDefinition("Recommendations", "Decision reports"),
    WorksheetDefinition("Advisor Report", "Decision reports"),
    WorksheetDefinition("Audit", "Validation and support"),
    WorksheetDefinition("Validation", "Validation and support"),
    WorksheetDefinition("Data Tables", "Validation and support"),
    WorksheetDefinition("Documentation", "Validation and support"),
    WorksheetDefinition("Change Log", "Validation and support"),
)

REQUIRED_WORKSHEET_ORDER: tuple[str, ...] = tuple(
    definition.name for definition in WORKSHEET_MANIFEST if definition.required
)

RESTORED_REPORTING_SHEETS: frozenset[str] = frozenset(
    {
        "Scenario Comparison", "Roth Conversion Optimizer", "Tax Optimizer",
        "Lifetime Cash Flow", "Sensitivity Analysis", "Monte Carlo Analysis",
        "ACA Analysis", "Medicare IRMAA", "Charitable Planning", "Estate Planning",
        "Healthcare Planning", "RMD Analysis", "Audit", "Validation",
        "Advisor Report", "Recommendations",
    }
)


def create_workbook_shell() -> Workbook:
    """Create an empty workbook with sheets in the approved user-workflow order."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    for definition in WORKSHEET_MANIFEST:
        workbook.create_sheet(definition.name)
    return workbook


def validate_sheet_structure(sheet_names: Sequence[str] | Iterable[str]) -> list[str]:
    """Return human-readable errors for missing, duplicate, or misordered sheets."""

    names = list(sheet_names)
    errors: list[str] = []
    duplicates = sorted({name for name in names if names.count(name) > 1})
    if duplicates:
        errors.append(f"Duplicate worksheets: {', '.join(duplicates)}")
    missing = [name for name in REQUIRED_WORKSHEET_ORDER if name not in names]
    if missing:
        errors.append(f"Missing required worksheets: {', '.join(missing)}")
    present_required = [name for name in names if name in REQUIRED_WORKSHEET_ORDER]
    expected_present_order = [name for name in REQUIRED_WORKSHEET_ORDER if name in present_required]
    if present_required != expected_present_order:
        errors.append("Required worksheets are not in the approved user-workflow order")
    return errors


def populate_rmd_analysis(workbook: Workbook, result: RmdAnnualProjectionResult) -> None:
    """Write auditable RMD/QCD projection results to the validation workbook."""

    result.validate_reconciliation()
    if "RMD Analysis" not in workbook.sheetnames:
        raise ValueError("workbook is missing the RMD Analysis worksheet")

    ws = workbook["RMD Analysis"]
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "RMD/QCD Analysis — Validation Candidate"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Projection Year"
    ws["B2"] = result.projection_year
    ws["A3"] = "Household Status"
    ws["B3"] = "COMPLIANT" if result.compliant else "ACTION REQUIRED"

    summary_rows = (
        ("Calculated RMD", result.household_calculated_rmd),
        ("Qualified QCD", result.household_qualified_qcd),
        ("Other Distributions", result.household_distributions_taken),
        ("Remaining Shortfall", result.household_remaining_shortfall),
    )
    for row, (label, value) in enumerate(summary_rows, start=5):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value).number_format = "$#,##0.00"

    headers = [
        "Owner", "Account", "Type", "Aggregation Group", "Prior-Year Balance",
        "Life Table", "Divisor", "Calculated RMD", "Qualified QCD",
        "Distributions", "Remaining Account RMD", "Violations",
    ]
    header_row = 11
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=column, value=header)
        cell.font = Font(bold=True)

    for row, account in enumerate(result.account_results, start=header_row + 1):
        values = (
            account.owner_id, account.account_id, account.account_type,
            account.aggregation_group, account.prior_year_end_balance,
            account.life_expectancy_table, account.divisor, account.calculated_rmd,
            account.qualified_qcd, account.distributions_taken,
            account.account_remaining_rmd, "; ".join(account.violations),
        )
        for column, value in enumerate(values, start=1):
            ws.cell(row=row, column=column, value=value)
        for column in (5, 8, 9, 10, 11):
            ws.cell(row=row, column=column).number_format = "$#,##0.00"

    validation = workbook["Validation"]
    validation["A1"] = "RMD/QCD Reconciliation"
    validation["A1"].font = Font(bold=True)
    validation["A2"] = "Status"
    validation["B2"] = "PASS" if result.compliant else "FAIL"
    validation["A3"] = "Remaining Shortfall"
    validation["B3"] = result.household_remaining_shortfall
    validation["B3"].number_format = "$#,##0.00"

    audit = workbook["Audit"]
    audit["A1"] = "RMD/QCD Audit Trail"
    audit["A1"].font = Font(bold=True)
    audit["A2"] = "Projection Year"
    audit["B2"] = result.projection_year
    audit["A3"] = "Account Count"
    audit["B3"] = len(result.account_results)
    audit["A4"] = "Aggregation Group Count"
    audit["B4"] = len(result.group_results)


def build_rmd_validation_candidate(
    result: RmdAnnualProjectionResult,
    output_path: str | Path,
) -> Path:
    """Create a separate RMD validation-candidate workbook and return its path."""

    path = Path(output_path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("validation candidate output must use the .xlsx extension")
    workbook = create_workbook_shell()
    populate_rmd_analysis(workbook, result)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
