"""Medicare IRMAA validation-candidate workbook and CSV reporting."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl.styles import Font

from retirement_planner.engines.irmaa_projection import IrmaaAnnualProjectionResult
from retirement_planner.reporting.excel_builder import create_workbook_shell


def populate_irmaa_reporting(workbook, result: IrmaaAnnualProjectionResult) -> None:
    """Populate IRMAA, audit, validation, advisor, and recommendation sheets."""

    result.validate_reconciliation()
    ws = workbook["Medicare IRMAA"]
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "Medicare IRMAA Analysis — Validation Candidate"
    ws["A1"].font = Font(bold=True, size=14)
    rows = (
        ("Premium Year", result.projection_year),
        ("Income Year", result.premium_result.income_year),
        ("Filing Status", result.premium_result.filing_status),
        ("IRMAA MAGI", result.premium_result.irmaa_magi),
        ("Tier", result.premium_result.tier),
        ("Next Threshold", result.premium_result.next_threshold),
        ("Distance to Next Threshold", result.premium_result.distance_to_next_threshold),
        ("Beneficiaries", result.premium_result.beneficiary_count),
        ("Annual Household Medicare Cost", result.annual_household_cost),
        ("Open Review Requests", result.open_review_count),
    )
    for row, (label, value) in enumerate(rows, start=3):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
    for row in (6, 8, 9, 11):
        ws.cell(row=row, column=2).number_format = "$#,##0.00"

    strategy = result.strategy_decision
    ws["A15"] = "Strategy Review"
    ws["A15"].font = Font(bold=True)
    if strategy is None:
        ws["B15"] = "No strategy tested"
    else:
        ws["A16"] = "Crossed Threshold"
        ws["B16"] = strategy.crossed_threshold
        ws["A17"] = "Allowed"
        ws["B17"] = strategy.allowed
        ws["A18"] = "Incremental Annual Cost"
        ws["B18"] = strategy.incremental_annual_cost
        ws["B18"].number_format = "$#,##0.00"
        ws["A19"] = "Warning"
        ws["B19"] = strategy.warning
        ws["A20"] = "Rationale"
        ws["B20"] = strategy.rationale

    diagnostic = result.income_diagnostic
    if diagnostic is not None:
        ws["A23"] = "Income Source"
        ws["B23"] = "Amount"
        ws["C23"] = "Category"
        ws["D23"] = "Tier Without"
        ws["E23"] = "Tier With"
        ws["F23"] = "Incremental Annual Cost"
        for cell in ws[23]:
            cell.font = Font(bold=True)
        for row, impact in enumerate(diagnostic.source_impacts, start=24):
            values = (
                impact.name,
                impact.amount,
                impact.category,
                impact.tier_without_source,
                impact.tier_with_source,
                impact.incremental_annual_cost,
            )
            for column, value in enumerate(values, start=1):
                ws.cell(row=row, column=column, value=value)
            ws.cell(row=row, column=2).number_format = "$#,##0.00"
            ws.cell(row=row, column=6).number_format = "$#,##0.00"

    validation = workbook["Validation"]
    validation["D1"] = "IRMAA Reconciliation"
    validation["D1"].font = Font(bold=True)
    validation["D2"] = "Status"
    validation["E2"] = "PASS"
    validation["D3"] = "Premium Year"
    validation["E3"] = result.projection_year

    audit = workbook["Audit"]
    audit["D1"] = "IRMAA Audit Trail"
    audit["D1"].font = Font(bold=True)
    audit["D2"] = "Tier"
    audit["E2"] = result.premium_result.tier
    audit["D3"] = "MAGI"
    audit["E3"] = result.premium_result.irmaa_magi
    audit["D4"] = "Review Requests"
    audit["E4"] = len(result.review_requests)

    advisor = workbook["Advisor Report"]
    advisor["A1"] = "Medicare IRMAA Advisor Summary"
    advisor["A1"].font = Font(bold=True)
    advisor["A2"] = (
        f"Premium-year {result.projection_year} IRMAA tier: "
        f"{result.premium_result.tier}; annual household Medicare cost: "
        f"${result.annual_household_cost:,.2f}."
    )
    if result.threshold_warning:
        advisor["A3"] = result.threshold_warning

    recommendations = workbook["Recommendations"]
    recommendations["A1"] = "Medicare IRMAA Recommendations"
    recommendations["A1"].font = Font(bold=True)
    if strategy is not None and strategy.crossed_threshold and not strategy.allowed:
        recommendations["A2"] = "Do not implement the proposed income strategy without explicit approval and documented rationale."
    elif result.premium_result.distance_to_next_threshold is not None:
        recommendations["A2"] = (
            "Maintain an IRMAA MAGI buffer of at least "
            f"${result.premium_result.distance_to_next_threshold:,.2f} below the next threshold."
        )
    else:
        recommendations["A2"] = "No higher IRMAA threshold remains in the current table."


def build_irmaa_validation_candidate(
    result: IrmaaAnnualProjectionResult,
    output_path: str | Path,
) -> Path:
    """Create a separate IRMAA validation-candidate workbook."""

    path = Path(output_path)
    if path.suffix.lower() != ".xlsx":
        raise ValueError("validation candidate output must use the .xlsx extension")
    workbook = create_workbook_shell()
    populate_irmaa_reporting(workbook, result)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def export_irmaa_audit_csv(
    result: IrmaaAnnualProjectionResult,
    output_path: str | Path,
) -> Path:
    """Export beneficiary/household IRMAA audit fields to CSV."""

    result.validate_reconciliation()
    path = Path(output_path)
    if path.suffix.lower() != ".csv":
        raise ValueError("IRMAA export must use the .csv extension")
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "premium_year", "income_year", "filing_status", "irmaa_magi", "tier",
        "beneficiary_count", "household_monthly_cost", "household_annual_cost",
        "next_threshold", "distance_to_next_threshold", "threshold_warning",
        "open_review_count",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerow(
            {
                "premium_year": result.projection_year,
                "income_year": result.premium_result.income_year,
                "filing_status": result.premium_result.filing_status,
                "irmaa_magi": result.premium_result.irmaa_magi,
                "tier": result.premium_result.tier,
                "beneficiary_count": result.premium_result.beneficiary_count,
                "household_monthly_cost": result.premium_result.household_monthly_cost,
                "household_annual_cost": result.premium_result.household_annual_cost,
                "next_threshold": result.premium_result.next_threshold,
                "distance_to_next_threshold": result.premium_result.distance_to_next_threshold,
                "threshold_warning": result.threshold_warning,
                "open_review_count": result.open_review_count,
            }
        )
    return path
