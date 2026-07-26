from pathlib import Path
from openpyxl import Workbook
ROOT=Path(__file__).resolve().parents[1]
OUTPUT=ROOT/'outputs'/'excel'/'retirement-planner.xlsx'
SHEETS=['Cover','Instructions','Household Inputs','Economic Assumptions','Asset-Class Assumptions','Tax Assumptions','Social Security','Pension and Other Income','Account Balances','RMD Calculations','Roth Conversion Strategy','Medicare IRMAA','ACA Premium Tax Credit','Healthcare Costs','Annual Projection','Scenario Comparison','Optimizer Results','Dashboard','Audit Checks','Data Tables','Documentation','Change Log']
OUTPUT.parent.mkdir(parents=True,exist_ok=True)
wb=Workbook(); wb.remove(wb.active)
for name in SHEETS:
    ws=wb.create_sheet(name); ws['A1']=name; ws['A2']='Placeholder structure — formulas and formatting to be added.'
wb.save(OUTPUT)
print(f'Created {OUTPUT}')
